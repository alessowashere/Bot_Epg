"""
Pipeline de resoluciones PDF 2026.

El flujo es intencionalmente conservador:
1. inventario: cuenta y clasifica nombres del ZIP.
2. extraer: lee texto de PDFs y genera CSV/JSONL revisable.
3. staging: guarda lo extraido en resoluciones_documentos.
4. importar: crea/actualiza expedientes solo desde staging OK.

Por defecto no modifica la BD. Los cambios reales requieren --aplicar.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
import unicodedata
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from database import Base, SessionLocal, engine
import models
from nombres import quitar_tratamientos
from identidad_academica import limpiar_programa_academico, normalizar_codigo_matricula
from requisitos_catalogo import inicializar_requisitos_expediente


ROOT = Path("/opt/sistema_posgrado")
INPUT_ROOT = ROOT / "data" / "input"
DEFAULT_ZIP = INPUT_ROOT / "resoluciones" / "2026" / "RESOLUCIONES FIRMADAS-20260707T150151Z-3-001.zip"
DEFAULT_OUT = ROOT / "data" / "resoluciones_2026"
DEFAULT_DOCENTES = INPUT_ROOT / "catalogos" / "DOCENTES.xlsx"
DEFAULT_RESOLUCIONES_EXCEL = INPUT_ROOT / "catalogos" / "LISTA DE RESOLUCIONES EMITIDAS 2025 (2).xlsx"

MESES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "setiembre": 9,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

PASOS = {
    1: ("Nombramiento de Asesor", ["nombramiento de asesor", "nombramiento asesor", "cambio de nombramiento de asesor", "cambio de asesor"]),
    2: ("Dictamen de Proyecto", ["dictamen de proyecto", "dictamen proyecto", "dictaminante de proyecto", "cambio de dictamen de proyecto", "cambio de dictaminante"]),
    3: ("Inscripcion del Proyecto", ["inscripcion de proyecto", "inscripción de proyecto", "inscrpcion de proyecto", "inscripcion de tema", "inscripción de tema", "ampliacion de inscripcion", "ampliación de inscripción"]),
    4: ("Declarado Apto", ["apto al grado", "declarar apto", "declarado apto", "-apto-", "apto para sustentacion"]),
    5: ("Dictamen de Tesis", ["dictamen de tesis", "dictaminante de tesis", "cambio de dictamen de tesis"]),
    6: ("Sustentacion", ["fecha y hora", "fijar fecha y hora", "sustentacion", "sustentación"]),
    7: ("Tramite del Diploma", ["diploma", "grado academico", "grado académico", "elevar al vrac"]),
}

TITULOS_DOCENTE = r"(?:Dr\.?|Dra\.?|Mtro\.?|Mtra\.?|Mtr\.?|Mg\.?|Mag\.?|Mgt\.?)"


@dataclass
class ResolucionExtraida:
    source_path: str
    source_hash: str
    archivo_normalizado: str
    resolucion_numero: str = ""
    resolucion_anio: int | None = None
    fecha_resolucion: str = ""
    expediente_admin: str = ""
    codigo_alumno: str = ""
    nombre_alumno: str = ""
    grado_postula: str = ""
    programa: str = ""
    titulo_tesis: str = ""
    tipo_resolucion: str = ""
    id_paso_inferido: int | None = None
    docentes_detectados: list[str] = field(default_factory=list)
    estado_revision: str = "OK"
    observaciones: list[str] = field(default_factory=list)
    texto_preview: str = ""


def clave_resolucion(numero) -> str:
    if numero in (None, ""):
        return ""
    texto = str(numero).strip()
    try:
        texto = str(int(float(texto)))
    except ValueError:
        match = re.search(r"\d+", texto)
        texto = match.group(0) if match else ""
    return texto.zfill(4) if texto else ""


def normalizar_ascii(texto: str) -> str:
    texto = texto or ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^A-Za-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip().upper()


def compactar(texto: str) -> str:
    return re.sub(r"\s+", " ", texto or "").strip()


def limitar_texto(texto: str | None, limite: int) -> str | None:
    if texto is None:
        return None
    texto = compactar(str(texto))
    if not texto:
        return None
    return texto[:limite]


def limpiar_nombre_persona(nombre: str) -> str:
    nombre = normalizar_ascii(quitar_tratamientos(nombre))
    nombre = re.sub(rf"^({TITULOS_DOCENTE.replace('?:', '')})\s+", "", nombre, flags=re.I)
    nombre = re.sub(r"\b(EL|LA|LOS|LAS|DOCENTES?|ESTUDIANTE|INTERESAD[AO]|BACH|BR|MGT|MG|MAG|MTR|MTRA|MTRO|DRA|DR)\b", " ", nombre)
    nombre = re.sub(r"\b(CAMBIO|DICTAMEN|DICTAMINANTE|RECTIFICACION|NOMBRAMIENTO|ASESOR|FECHA|HORA|TESIS|PROYECTO|UAC|EPG)\b", " ", nombre)
    nombre = re.sub(r"\bDE\s*MATR(?:ICULA)?\b.*$", " ", nombre)
    nombre = re.sub(r"\bDE\s*$", " ", nombre)
    return re.sub(r"\s+", " ", nombre).strip()


def slug_archivo(nombre: str, limite: int = 180) -> str:
    base = Path(nombre).stem
    base = normalizar_ascii(base)
    base = re.sub(r"\bRESOLUCION\b", "RES", base)
    base = re.sub(r"\bEPG UAC\b", "EPG-UAC", base)
    base = base[:limite].strip()
    return f"{base}.pdf"


def hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def iter_pdfs_zip(zip_path: Path):
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir() or not info.filename.lower().endswith(".pdf"):
                continue
            yield info, zf.read(info.filename)


@dataclass
class ArchivoLocal:
    """Adaptador mínimo para reutilizar el extractor ZIP con archivos de Drive."""
    filename: str


def iter_pdfs_directorio(directorio: Path):
    """Lee PDFs recursivamente sin modificar los originales descargados."""
    for ruta in sorted(directorio.rglob("*.pdf")):
        if ruta.is_file():
            yield ArchivoLocal(str(ruta.relative_to(directorio))), ruta.read_bytes()


def extraer_texto_pdf(data: bytes, max_paginas: int) -> str:
    texto = []
    with pdfplumber.open(BytesIO(data)) as pdf:
        for page in pdf.pages[:max_paginas]:
            texto.append(page.extract_text() or "")
    return compactar("\n".join(texto))


def detectar_resolucion(texto: str, archivo: str) -> tuple[str, int | None]:
    patrones = [
        r"RESOLUCI[OÓ]N\s*[NW]\s*[\s.°º]*0*([0-9]{1,4})\s*[-/·–— ]+\s*(20[0-9]{2})",
        r"\bN\s*[\s.°º]*0*([0-9]{1,4})\s*[-/·–— ]+\s*(20[0-9]{2})",
    ]
    # La cabecera debe prevalecer sobre resoluciones citadas en antecedentes.
    for origen in ((texto or "")[:1800], archivo):
        for patron in patrones:
            match = re.search(patron, origen or "", re.I)
            if match:
                return match.group(1).zfill(4), int(match.group(2))
    return "", None


def detectar_fecha(texto: str) -> str:
    texto_fecha = re.sub(r"(?<=\d)(?=[A-Za-zÁÉÍÓÚáéíóú])", " ", texto or "")
    texto_fecha = re.sub(r"(?<=[A-Za-zÁÉÍÓÚáéíóú])(?=\d)", " ", texto_fecha)
    texto_fecha = texto_fecha.replace("d e", "de")
    meses = "|".join(MESES)
    texto_fecha = re.sub(rf"de\s*({meses})\s*(?:de|del)?\s*(20\d{{2}})", r"de \1 de \2", texto_fecha, flags=re.I)
    for mes in MESES:
        patron_espaciado = r"\s*".join(re.escape(letra) for letra in mes)
        texto_fecha = re.sub(patron_espaciado, mes, texto_fecha, flags=re.I)
    match = re.search(
        r"Cus(?:co|ca|zco)\s*[,.:]?\s*(\d{1,2})\s*(?:de\s*)?([a-záéíóú]+)\s*(?:de\s*)?(?:del\s*)?(20\d{2})",
        texto_fecha,
        re.I,
    )
    if not match:
        return ""
    dia = int(match.group(1))
    mes = MESES.get(normalizar_ascii(match.group(2)).lower())
    anio = int(match.group(3))
    if not mes:
        return ""
    try:
        return datetime(anio, mes, dia).date().isoformat()
    except ValueError:
        return ""


def detectar_paso_y_tipo(texto: str, archivo: str, tipo_excel: str = "") -> tuple[int | None, str]:
    objetivo_excel = normalizar_ascii(tipo_excel)
    objetivo_archivo = normalizar_ascii(archivo)
    objetivo_texto = normalizar_ascii(texto[:6000])
    prioridad = [1, 2, 3, 5, 6, 4, 7]
    for paso in prioridad:
        nombre, claves = PASOS[paso]
        if any(normalizar_ascii(clave) in objetivo_excel for clave in claves):
            return paso, nombre
    for paso in prioridad:
        nombre, claves = PASOS[paso]
        if any(normalizar_ascii(clave) in objetivo_archivo for clave in claves):
            return paso, nombre
    for paso in prioridad:
        nombre, claves = PASOS[paso]
        if any(normalizar_ascii(clave) in objetivo_texto for clave in claves):
            return paso, nombre
    return None, ""


def detectar_grado(texto: str, archivo: str) -> str:
    objetivo = normalizar_ascii(f"{texto[:7000]} {archivo}")
    match = re.search(r"(?:PARA\s*OPTAR|ASPIRANTE)\s*AL\s*GRADO\s*ACADEMICO\s*DE\s*([^.;]+)", objetivo)
    fragmento = match.group(1) if match else objetivo
    if re.search(r"\b(DOCTOR|DOCTORA|DOCTORADO)\b", fragmento):
        return "Doctor"
    if re.search(r"\b(MAESTRO|MAESTRA|MAESTRIA|MAGISTER)\b", fragmento):
        return "Maestro"
    if re.search(r"PROGRAMA\s*DE\s*DOCTORADO", objetivo):
        return "Doctor"
    if re.search(r"PROGRAMA\s*DE\s*MAESTRIA", objetivo):
        return "Maestro"
    return ""


def detectar_programa(texto: str) -> str:
    patrones = [
        r"programa\s*de\s*((?:MAESTR[IÍ]A|DOCTORADO)\s*en\s*.*?)(?:de\s*la\s*Escuela|\s*;|\.|\n)",
        r"programa\s*de\s*((?:MAESTR[IÍ]A|DOCTORADO).*?)(?:\s+de la Escuela|\s*;|\.|\n)",
        r"Grado Acad[eé]mico de\s+((?:MAESTR[OA]|DOCTOR[AO]).*?)(?:\s+habiendo|\s+de la Escuela|\s*;|\.|\n)",
    ]
    for patron in patrones:
        match = re.search(patron, texto, re.I)
        if match:
            programa = limpiar_programa_academico(match.group(1))
            programa = re.split(
                r"\s*,?\s+DE LA UNIVERSIDAD\b|\s*,?\s+QUIEN SOLICITA\b|\s*,?\s+SOLICITAD[OA] POR\b|\s*,?\s+CUYAS FUNCIONES\b|\s*,?\s+DE CONFORMIDAD\b|\s*,?\s+DE ACUERDO\b|\s*,?\s+RESULTA\b|\s*,?\s+TRAS LA\b|DE LA ESCUELA\b|\s+DEL INTERESAD[OA]\b|\s+DEL BACH\b|\s+DE LA BR\b|\s+DEL BR\b|\s+DE LA MTR[AO]?\b|\s+DEL MTR[AO]?\b",
                programa,
                maxsplit=1,
                flags=re.I,
            )[0]
            return limitar_texto(programa, 250) or ""
    return ""


def detectar_titulo(texto: str) -> str:
    patrones = [
        r"tesis intitulada:\s*[\"“]([^\"”]+)[\"”]",
        r"Tesis Intitulada:\s*[\"“]([^\"”]+)[\"”]",
        r"tema de tesis intitulada:\s*[\"“]([^\"”]+)[\"”]",
    ]
    encontrados = []
    for patron in patrones:
        for match in re.finditer(patron, texto, re.I):
            encontrados.append((match.start(), compactar(match.group(1)).upper()))
    if not encontrados:
        return ""
    return sorted(encontrados, key=lambda item: item[0])[-1][1]


def cargar_catalogo_excel_2026(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    if "RESOLUCIONES 2026" not in wb.sheetnames:
        return {}
    ws = wb["RESOLUCIONES 2026"]
    catalogo = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        numero = clave_resolucion(row[1] if len(row) > 1 else None)
        if not numero:
            continue
        tipo = compactar(str(row[5] or "")) if len(row) > 5 else ""
        paso, tipo_normalizado = detectar_paso_y_tipo("", "", tipo)
        catalogo[numero] = {
            "estado_excel": compactar(str(row[0] or "")) if len(row) > 0 else "",
            "resolucion_numero": numero,
            "fecha_excel": row[2] if len(row) > 2 else None,
            "nombre_excel": limpiar_nombre_persona(str(row[3] or "")) if len(row) > 3 and row[3] else "",
            "codigo_excel": normalizar_codigo_matricula(row[4]) if len(row) > 4 and row[4] else "",
            "tipo_excel": tipo,
            "id_paso_excel": paso,
            "tipo_paso_excel": tipo_normalizado,
            "observacion_excel": compactar(str(row[6] or "")) if len(row) > 6 else "",
        }
    return catalogo


def estado_excel_excluye(estado: str) -> bool:
    estado_norm = normalizar_ascii(estado)
    return "DEJAR SIN EFECTO" in estado_norm or estado_norm == "NO ENVIADO"


def excel_row_compatible(excel_row: dict | None, nombre: str, codigo: str) -> bool:
    if not excel_row:
        return False
    codigo_excel = excel_row.get("codigo_excel") or ""
    nombre_excel = normalizar_ascii(excel_row.get("nombre_excel") or "")
    nombre_pdf = normalizar_ascii(nombre or "")
    if codigo and codigo_excel and codigo != codigo_excel:
        return False
    if nombre_pdf and nombre_excel and nombre_excel not in nombre_pdf and nombre_pdf not in nombre_excel:
        return False
    return True


def fecha_excel_iso(excel_row: dict | None) -> str:
    if not excel_row:
        return ""
    fecha = excel_row.get("fecha_excel")
    if isinstance(fecha, datetime):
        return fecha.date().isoformat()
    return ""


def resoluciones_referidas(texto: str, numero_actual: str = "") -> list[str]:
    refs = []
    for match in re.finditer(r"RESOLUCI[OÓ]N\s*N\s*[\s.°º]*0*([0-9]{1,5})\s*[-/ ]+\s*(20[0-9]{2})", texto or "", re.I):
        numero = match.group(1).zfill(4)
        if numero and numero != numero_actual and numero not in refs:
            refs.append(numero)
    return refs


def detectar_alumno_codigo(texto: str, archivo: str) -> tuple[str, str, str]:
    codigo = ""
    codigo_match = re.search(
        r"c[oó]digo(?:\s+de\s+matr[ií]cula)?[,]?\s*(?:N\s*[\s.°º]*)?([A-Z0-9]{8,16})",
        texto,
        re.I,
    )
    if codigo_match:
        codigo = normalizar_codigo_matricula(codigo_match.group(1))

    expediente = ""
    exp_match = re.search(r"expediente\s*(?:administrativo)?\s*(?:N[°º.]?\s*)?#?\s*([0-9]+)", texto, re.I)
    if exp_match:
        expediente = exp_match.group(1)

    nombre = ""
    bloque_visto = texto
    match_visto = re.search(r"\bVISTO\b(?:\s*[:.-])?\s*(.*?)(?:\bCONSIDERANDO\b|\bCONSIDERANDO\.|\bQue,)", texto, re.I)
    if match_visto:
        bloque_visto = match_visto.group(1)
    patrones = [
        r"presentad[oa]\s*por\s*(?:el|la)?\s*(?:estudiante\s*)?(?:Bach\.?|Br\.?|Mgt\.?|Mg\.?|Mag\.?|Mtr\.?|Mtra\.?|Mtro\.?|Dr\.?|Dra\.?|Sr\.?|Sra\.?|Don|Doña)?\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s]+?)(?:,?\s*con\s*c[oó]digo|,?\s+y\s+el\b|,?\s+quien\b|,?\s+participante\b|,?\s+en\s+representaci[oó]n\b|;|\\.)",
        r"presentad[oa]\s+por\s+(?:el|la)?\s*estudiante\s+(?:Bach\.?|Br\.?|Mgt\.?|Mg\.?|Mag\.?|Mtr\.?|Mtra\.?|Mtro\.?)?\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)(?:,|\s+con\s+c[oó]digo|\\.)",
        r"correspondiente\s+a\s+(?:el|la)?\s*(?:Bach\.?|Br\.?|Mgt\.?|Mg\.?|Mag\.?|Mtr\.?|Mtra\.?|Mtro\.?|Dr\.?|Dra\.?)?\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)(?:,|\s+y\b|;|\\.)",
        r"presentada por\s+(?:el|la)\s+estudiante\s+(?:Bach\.?|Br\.?|Mgt\.?|Mg\.?|Mag\.?|Mtr\.?|Mtra\.?|Mtro\.?)?\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?),?\s+para optar",
    ]
    for patron in patrones:
        match = re.search(patron, bloque_visto, re.I)
        if match:
            nombre = limpiar_nombre_persona(match.group(1))
            break
    return nombre, codigo, expediente


def detectar_docentes(texto: str) -> list[str]:
    docentes = set()
    segmentos = []
    for patron in [
        r"(?:docentes|docente|dictaminantes?|asesor(?:a)?(?:es)?)[^.;:]{0,120}",
        r"(?:ENCOMENDAR|NOMBRAR|DESIGNAR)[^.;]{0,260}",
    ]:
        segmentos.extend(m.group(0) for m in re.finditer(patron, texto, re.I))

    for segmento in segmentos:
        for match in re.finditer(rf"{TITULOS_DOCENTE}\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ,\s]{{6,90}})", segmento, re.I):
            nombre = match.group(1)
            nombre = re.split(r"\s+(?:y|como|en calidad|del proyecto|de la tesis)\s+", nombre, maxsplit=1, flags=re.I)[0]
            nombre = nombre.replace(",", " ")
            limpio = limpiar_nombre_persona(nombre)
            if len(limpio.split()) >= 2:
                docentes.add(limpio)
    return sorted(docentes)


def evaluar_observaciones(item: ResolucionExtraida, texto: str, excel_row: dict | None = None) -> None:
    if not item.resolucion_numero:
        item.observaciones.append("sin_numero_resolucion")
    if not item.fecha_resolucion:
        item.observaciones.append("sin_fecha")
    if not item.nombre_alumno:
        item.observaciones.append("sin_nombre_alumno")
    if not item.codigo_alumno:
        item.observaciones.append("sin_codigo_alumno")
    if not item.grado_postula:
        item.observaciones.append("sin_grado")
    if not item.id_paso_inferido:
        item.observaciones.append("sin_paso")
    if excel_row:
        estado_excel = excel_row.get("estado_excel") or ""
        if estado_excel and normalizar_ascii(estado_excel) != "ENVIADO":
            item.observaciones.append(f"estado_excel_{normalizar_ascii(estado_excel).lower().replace(' ', '_')}")
    if item.observaciones:
        item.estado_revision = "Observado"


def es_documento_resolucion(texto: str, archivo: str) -> bool:
    """Evita que actas, oficios y compilados entren por mencionar un paso."""
    inicio = normalizar_ascii((texto or "")[:1600])
    nombre_archivo = normalizar_ascii(archivo or "")
    encabezado = re.search(
        r"\bRESOLUCION\s*[NW](?:RO)?\s*[.°º ]*0*[0-9]{1,4}\s*[-/ ]+\s*20[0-9]{2}",
        inicio,
        re.I,
    )
    documento_ajeno = re.search(
        r"\b(?:ACTA\s+DE\s+(?:SUSTENTACION|SESION)|OFICIO|INFORME|MEMORANDUM)\b",
        inicio[:700],
        re.I,
    )
    if documento_ajeno and (not encabezado or documento_ajeno.start() < encabezado.start()):
        return False
    if encabezado:
        return True
    return bool(
        re.search(r"RESOLUCION", nombre_archivo, re.I)
        and re.search(r"0*[0-9]{1,5}\s*[-_/ ]+\s*20[0-9]{2}", nombre_archivo)
    )


def extraer_resolucion(info, data: bytes, max_paginas: int, catalogo_excel: dict[str, dict] | None = None) -> ResolucionExtraida | None:
    source_hash = hash_bytes(data)
    archivo = info.filename
    texto = ""
    observaciones = []
    try:
        texto = extraer_texto_pdf(data, max_paginas=max_paginas)
    except Exception as exc:
        observaciones.append(f"error_pdf:{exc}")

    if not es_documento_resolucion(texto, archivo):
        return None

    numero, anio = detectar_resolucion(texto, archivo)
    nombre, codigo, expediente = detectar_alumno_codigo(texto, archivo)
    excel_row = (catalogo_excel or {}).get(numero)
    excel_row = excel_row if excel_row_compatible(excel_row, nombre, codigo) else None
    if excel_row and estado_excel_excluye(excel_row.get("estado_excel", "")):
        return None

    tipo_excel = excel_row.get("tipo_excel", "") if excel_row else ""
    paso, tipo = detectar_paso_y_tipo(texto, archivo, tipo_excel)
    if not paso:
        return None

    if excel_row:
        nombre = nombre or excel_row.get("nombre_excel", "")
        codigo = codigo or excel_row.get("codigo_excel", "")
        if excel_row.get("id_paso_excel"):
            paso = excel_row["id_paso_excel"]
            tipo = excel_row["tipo_paso_excel"]
    if catalogo_excel and (not nombre or not codigo):
        for ref in resoluciones_referidas(texto, numero):
            ref_row = catalogo_excel.get(ref)
            if not ref_row:
                continue
            nombre = nombre or ref_row.get("nombre_excel", "")
            codigo = codigo or ref_row.get("codigo_excel", "")
            if nombre and codigo:
                break
    item = ResolucionExtraida(
        source_path=archivo,
        source_hash=source_hash,
        archivo_normalizado=slug_archivo(archivo),
        resolucion_numero=numero,
        resolucion_anio=anio,
        fecha_resolucion=detectar_fecha(texto) or fecha_excel_iso(excel_row),
        expediente_admin=expediente,
        codigo_alumno=codigo,
        nombre_alumno=nombre,
        grado_postula=detectar_grado(texto, archivo),
        programa=detectar_programa(texto),
        titulo_tesis=detectar_titulo(texto),
        tipo_resolucion=tipo,
        id_paso_inferido=paso,
        docentes_detectados=detectar_docentes(texto),
        observaciones=observaciones,
        texto_preview=texto[:3000],
    )
    evaluar_observaciones(item, texto, excel_row)
    return item


def completar_por_codigo(items: list[ResolucionExtraida]) -> None:
    por_codigo = {}
    for item in items:
        if item.codigo_alumno:
            actual = por_codigo.setdefault(item.codigo_alumno, {})
            for campo in ("grado_postula", "programa", "titulo_tesis", "nombre_alumno"):
                valor = getattr(item, campo)
                if valor and not actual.get(campo):
                    actual[campo] = valor

    for item in items:
        if not item.codigo_alumno:
            continue
        fuente = por_codigo.get(item.codigo_alumno, {})
        cambio = False
        for campo in ("grado_postula", "programa", "titulo_tesis", "nombre_alumno"):
            if not getattr(item, campo) and fuente.get(campo):
                setattr(item, campo, fuente[campo])
                cambio = True
        if cambio:
            item.observaciones = [
                obs for obs in item.observaciones
                if obs not in {"sin_grado", "sin_nombre_alumno"}
            ]
            if item.observaciones:
                item.estado_revision = "Observado"
            else:
                item.estado_revision = "OK"


def escribir_csv(items: list[ResolucionExtraida], ruta: Path) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    campos = list(asdict(items[0]).keys()) if items else list(ResolucionExtraida("", "", "").__dict__.keys())
    with ruta.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=campos)
        writer.writeheader()
        for item in items:
            row = asdict(item)
            row["docentes_detectados"] = " | ".join(item.docentes_detectados)
            row["observaciones"] = " | ".join(item.observaciones)
            writer.writerow(row)


def escribir_jsonl(items: list[ResolucionExtraida], ruta: Path) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with ruta.open("w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(asdict(item), ensure_ascii=False) + "\n")


def comando_inventario(args) -> None:
    zip_path = Path(args.zip)
    out = Path(args.out)
    conteo = {}
    filas = []
    for info, data in iter_pdfs_zip(zip_path):
        paso, tipo = detectar_paso_y_tipo("", info.filename)
        numero, anio = detectar_resolucion("", info.filename)
        conteo[tipo or "Sin clasificar"] = conteo.get(tipo or "Sin clasificar", 0) + 1
        filas.append({
            "source_path": info.filename,
            "bytes": info.file_size,
            "source_hash": hash_bytes(data),
            "resolucion_numero": numero,
            "resolucion_anio": anio,
            "tipo_nombre_archivo": tipo,
            "id_paso_nombre_archivo": paso,
            "archivo_normalizado": slug_archivo(info.filename),
        })

    out.mkdir(parents=True, exist_ok=True)
    csv_path = out / "inventario_zip.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(filas[0].keys()) if filas else [])
        writer.writeheader()
        writer.writerows(filas)
    print(f"PDFs encontrados: {len(filas)}")
    for tipo, total in sorted(conteo.items()):
        print(f"  {tipo}: {total}")
    print(f"Inventario: {csv_path}")


def extraer_items(items_origen, args) -> None:
    out = Path(args.out)
    limite = args.limite
    catalogo_excel = cargar_catalogo_excel_2026(Path(args.excel_resoluciones)) if args.excel_resoluciones else {}
    items = []
    excluidos = 0
    for idx, (info, data) in enumerate(items_origen, start=1):
        if limite and idx > limite:
            break
        if idx % 25 == 0:
            print(f"Procesando {idx} PDFs...")
        item = extraer_resolucion(info, data, max_paginas=args.paginas, catalogo_excel=catalogo_excel)
        if item is None:
            excluidos += 1
            continue
        items.append(item)

    completar_por_codigo(items)

    csv_path = out / "resoluciones_extraidas.csv"
    jsonl_path = out / "resoluciones_extraidas.jsonl"
    escribir_csv(items, csv_path)
    escribir_jsonl(items, jsonl_path)
    total_obs = sum(1 for item in items if item.estado_revision == "Observado")
    print(f"Extraidos: {len(items)}")
    print(f"Excluidos: {excluidos}")
    print(f"Observados: {total_obs}")
    print(f"CSV: {csv_path}")
    print(f"JSONL: {jsonl_path}")


def comando_extraer(args) -> None:
    extraer_items(iter_pdfs_zip(Path(args.zip)), args)


def comando_extraer_directorio(args) -> None:
    directorio = Path(args.directorio)
    if not directorio.is_dir():
        raise FileNotFoundError(f"No existe el directorio de PDFs: {directorio}")
    extraer_items(iter_pdfs_directorio(directorio), args)


def cargar_jsonl(ruta: Path) -> list[dict]:
    items = []
    with ruta.open(encoding="utf-8") as f:
        for line in f:
            if line.strip():
                items.append(json.loads(line))
    return items


def parse_fecha(fecha: str):
    if not fecha:
        return None
    return datetime.strptime(fecha, "%Y-%m-%d")


def fecha_vencimiento_resolucion(fecha: datetime | None) -> datetime | None:
    """Toda resolución de los siete pasos tiene vigencia de 24 meses.

    `replace` conserva el día/mes y resuelve correctamente el caso excepcional
    del 29 de febrero al 28 de febrero del segundo año.
    """
    if not fecha:
        return None
    try:
        return fecha.replace(year=fecha.year + 2)
    except ValueError:
        return fecha.replace(year=fecha.year + 2, month=2, day=28)


def actualizar_vigencia_expediente(db, expediente) -> bool:
    """Marca vencido solo el paso vigente; P7 queda como expediente concluido."""
    if expediente.id_paso_actual == 7:
        expediente.estado_expediente = "Archivado_Graduado"
        expediente.sub_estado = "Concluido con resolución de grado"
        return False

    ultima = (
        db.query(models.ResolucionFirma)
        .filter(
            models.ResolucionFirma.id_expediente == expediente.id_expediente,
            models.ResolucionFirma.id_paso_asociado == expediente.id_paso_actual,
            models.ResolucionFirma.estado_firma == "Firmado",
        )
        .order_by(models.ResolucionFirma.fecha_firma.desc())
        .first()
    )
    vencimiento = fecha_vencimiento_resolucion(ultima.fecha_firma if ultima else None)
    if vencimiento and datetime.utcnow() >= vencimiento:
        expediente.estado_expediente = "Caduco"
        expediente.sub_estado = f"Paso {expediente.id_paso_actual} vencido el {vencimiento.date().isoformat()}"
        return True
    if expediente.estado_expediente == "Caduco":
        expediente.estado_expediente = "En Proceso"
        expediente.sub_estado = None
    return False


def comando_staging(args) -> None:
    Base.metadata.create_all(bind=engine)
    items = cargar_jsonl(Path(args.jsonl))
    db = SessionLocal()
    nuevos = 0
    actualizados = 0
    duplicados_lote = 0
    hashes_lote = set()
    try:
        for row in items:
            source_hash = row["source_hash"]
            if source_hash in hashes_lote:
                duplicados_lote += 1
                continue
            hashes_lote.add(source_hash)
            existente = db.query(models.ResolucionDocumento).filter(
                models.ResolucionDocumento.source_hash == source_hash
            ).first()
            estado = "OK" if row.get("estado_revision") == "OK" else "Observado"
            payload = {
                "source_path": limitar_texto(row.get("source_path"), 500),
                "archivo_normalizado": limitar_texto(row.get("archivo_normalizado"), 255),
                "resolucion_numero": limitar_texto(row.get("resolucion_numero"), 30),
                "resolucion_anio": row.get("resolucion_anio"),
                "fecha_resolucion": parse_fecha(row.get("fecha_resolucion") or ""),
                "expediente_admin": limitar_texto(row.get("expediente_admin"), 30),
                "codigo_alumno": limitar_texto(row.get("codigo_alumno"), 30),
                "nombre_alumno": limitar_texto(row.get("nombre_alumno"), 200),
                "grado_postula": row.get("grado_postula") or None,
                "programa": limitar_texto(detectar_programa(row.get("texto_preview") or "") or row.get("programa"), 250),
                "titulo_tesis": row.get("titulo_tesis") or None,
                "tipo_resolucion": limitar_texto(row.get("tipo_resolucion"), 120),
                "id_paso_inferido": row.get("id_paso_inferido"),
                "docentes_detectados": row.get("docentes_detectados") or [],
                "texto_preview": row.get("texto_preview") or None,
                "estado_revision": estado,
                "observaciones": " | ".join(row.get("observaciones") or []),
            }
            if existente:
                for key, value in payload.items():
                    setattr(existente, key, value)
                actualizados += 1
            else:
                db.add(models.ResolucionDocumento(source_hash=source_hash, **payload))
                nuevos += 1
        if args.aplicar:
            db.commit()
        else:
            db.rollback()
        print(
            f"Staging {'aplicado' if args.aplicar else 'simulado'}: "
            f"{nuevos} nuevos, {actualizados} actualizados, {duplicados_lote} duplicados en lote omitidos"
        )
    finally:
        db.close()


def detectar_columnas_docentes(sheet):
    headers = [str(c.value or "").strip().lower() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    col_nombre = next((i for i, h in enumerate(headers) if "apellidos" in h or "nombre" in h), None)
    col_dni = next((i for i, h in enumerate(headers) if "documento" in h or "dni" in h), None)
    col_correo = next((i for i, h in enumerate(headers) if "correo" in h or "email" in h), None)
    return col_nombre, col_dni, col_correo


def comando_importar_docentes(args) -> None:
    wb = load_workbook(args.excel, data_only=True, read_only=True)
    sheet = wb.active
    col_nombre, col_dni, col_correo = detectar_columnas_docentes(sheet)
    if col_nombre is None:
        raise RuntimeError("No se encontro columna de nombre en DOCENTES.xlsx")
    db = SessionLocal()
    nuevos = 0
    actualizados = 0
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre = limpiar_nombre_persona(str(row[col_nombre] or ""))
            if not nombre or nombre == "APELLIDOS Y NOMBRES":
                continue
            dni = str(row[col_dni] or "").strip() if col_dni is not None else None
            correo = str(row[col_correo] or "").strip().lower() if col_correo is not None and row[col_correo] else None
            docente = db.query(models.Docente).filter(models.Docente.dni == dni).first() if dni else None
            if not docente:
                docente = db.query(models.Docente).filter(models.Docente.nombre_completo == nombre).first()
            if docente:
                docente.nombre_completo = nombre
                docente.correo = correo or docente.correo
                actualizados += 1
            else:
                db.add(models.Docente(
                    dni=dni or None,
                    nombre_completo=nombre,
                    correo=correo,
                    tipo_contrato="Semestral",
                    estado="Activo",
                    max_tesis_permitidas=5,
                ))
                nuevos += 1
        if args.aplicar:
            db.commit()
        else:
            db.rollback()
        print(f"Docentes {'aplicados' if args.aplicar else 'simulados'}: {nuevos} nuevos, {actualizados} actualizados")
    finally:
        db.close()


def comando_importar_expedientes(args) -> None:
    if args.aplicar:
        print(
            "Importación omitida de forma segura: la agrupación histórica por código no es segura. "
            "Primero generar y revisar el catálogo de identidades académicas; usar después "
            "el reconstruidor compuesto, no este importador legado."
        )
        return
    db = SessionLocal()
    creados = 0
    actualizados = 0
    observados = 0
    resoluciones = 0
    caducos = 0
    expedientes_tocados = {}
    try:
        query = db.query(models.ResolucionDocumento).filter(models.ResolucionDocumento.estado_revision == "OK")
        for doc in query.order_by(models.ResolucionDocumento.fecha_resolucion, models.ResolucionDocumento.resolucion_numero).all():
            if not doc.codigo_alumno or not doc.nombre_alumno or not doc.grado_postula or not doc.id_paso_inferido:
                observados += 1
                doc.estado_revision = "Observado"
                doc.observaciones = compactar(f"{doc.observaciones or ''} | incompleto_para_importar")
                continue

            expediente = db.query(models.ExpedienteTesis).filter(
                models.ExpedienteTesis.codigo_alumno == doc.codigo_alumno
            ).first()
            if not expediente:
                expediente = models.ExpedienteTesis(
                    codigo_alumno=doc.codigo_alumno,
                    nombre_alumno=doc.nombre_alumno,
                    grado_postula=doc.grado_postula,
                    programa=doc.programa or None,
                    titulo_tesis=doc.titulo_tesis,
                    id_paso_actual=doc.id_paso_inferido,
                    estado_expediente="En Proceso",
                    fecha_inicio_tramite=doc.fecha_resolucion or datetime.utcnow(),
                )
                db.add(expediente)
                db.flush()
                inicializar_requisitos_expediente(db, expediente)
                db.add(models.HistorialMovimiento(
                    id_expediente=expediente.id_expediente,
                    id_paso=doc.id_paso_inferido,
                    accion="Creado",
                    nota=f"Creado desde resolucion {doc.resolucion_numero}-{doc.resolucion_anio}",
                    usuario_nombre="Sistema (Resoluciones PDF)",
                ))
                creados += 1
            else:
                expediente.nombre_alumno = doc.nombre_alumno
                expediente.grado_postula = doc.grado_postula
                if doc.programa:
                    expediente.programa = doc.programa
                if doc.titulo_tesis:
                    expediente.titulo_tesis = doc.titulo_tesis
                if doc.id_paso_inferido and doc.id_paso_inferido > expediente.id_paso_actual:
                    expediente.id_paso_actual = doc.id_paso_inferido
                actualizados += 1
            expedientes_tocados[expediente.id_expediente] = expediente

            etiqueta = f"Resolucion {doc.resolucion_numero}-{doc.resolucion_anio} - {doc.tipo_resolucion}"[:100]
            existe = db.query(models.ResolucionFirma).filter(
                models.ResolucionFirma.id_expediente == expediente.id_expediente,
                models.ResolucionFirma.tipo_documento == etiqueta,
            ).first()
            if not existe:
                db.add(models.ResolucionFirma(
                    id_expediente=expediente.id_expediente,
                    id_paso_asociado=doc.id_paso_inferido,
                    tipo_documento=etiqueta[:100],
                    archivo_drive_url=doc.source_path,
                    estado_firma="Firmado",
                    fecha_firma=doc.fecha_resolucion or datetime.utcnow(),
                ))
                db.add(models.HistorialMovimiento(
                    id_expediente=expediente.id_expediente,
                    id_paso=doc.id_paso_inferido,
                    accion="Resolucion_Cargada",
                    nota=f"{etiqueta} | fuente: {doc.source_path}",
                    usuario_nombre="Sistema (Resoluciones PDF)",
                    fecha_movimiento=doc.fecha_resolucion or datetime.utcnow(),
                ))
                resoluciones += 1

            doc.estado_revision = "Importado"

        for expediente in expedientes_tocados.values():
            caducos += 1 if actualizar_vigencia_expediente(db, expediente) else 0

        if args.aplicar:
            db.commit()
        else:
            db.rollback()
        print(
            f"Expedientes {'aplicados' if args.aplicar else 'simulados'}: "
            f"{creados} creados, {actualizados} actualizados, {resoluciones} resoluciones, {observados} observados, {caducos} caducos"
        )
    finally:
        db.close()


def comando_ordenar_archivos(args) -> None:
    zip_path = Path(args.zip)
    destino = Path(args.destino)
    if destino.exists() and args.aplicar:
        raise RuntimeError(f"El destino ya existe: {destino}. Usa otro destino para no mezclar archivos.")
    total = 0
    for info, data in iter_pdfs_zip(zip_path):
        numero, anio = detectar_resolucion("", info.filename)
        paso, tipo = detectar_paso_y_tipo("", info.filename)
        carpeta = f"{paso or 0:02d}_{normalizar_ascii(tipo or 'SIN_CLASIFICAR').replace(' ', '_')}"
        nombre = slug_archivo(info.filename)
        if numero and anio:
            nombre = f"{anio}-{numero}_{nombre}"
        salida = destino / carpeta / nombre
        total += 1
        if args.aplicar:
            salida.parent.mkdir(parents=True, exist_ok=True)
            salida.write_bytes(data)
    print(f"Archivos {'ordenados' if args.aplicar else 'simulados'}: {total}")
    print(f"Destino: {destino}")


def comando_validar_secuencia(args) -> None:
    items = cargar_jsonl(Path(args.jsonl))
    por_codigo = {}
    sin_codigo = 0
    for item in items:
        codigo = item.get("codigo_alumno")
        paso = item.get("id_paso_inferido")
        if not codigo:
            sin_codigo += 1
            continue
        data = por_codigo.setdefault(codigo, {
            "codigo_alumno": codigo,
            "nombre_alumno": item.get("nombre_alumno") or "",
            "grado_postula": item.get("grado_postula") or "",
            "pasos": set(),
            "resoluciones": [],
        })
        if paso:
            data["pasos"].add(int(paso))
        if item.get("resolucion_numero"):
            data["resoluciones"].append(f"{item.get('resolucion_numero')}-{item.get('resolucion_anio')}")

    filas = []
    for data in por_codigo.values():
        pasos = sorted(data["pasos"])
        max_paso = max(pasos) if pasos else 0
        faltantes = [p for p in range(1, max_paso) if p not in data["pasos"]]
        estado = "Observado" if faltantes else "OK"
        filas.append({
            "codigo_alumno": data["codigo_alumno"],
            "nombre_alumno": data["nombre_alumno"],
            "grado_postula": data["grado_postula"],
            "pasos_detectados": " ".join(str(p) for p in pasos),
            "max_paso": max_paso,
            "pasos_faltantes_previos": " ".join(str(p) for p in faltantes),
            "estado_secuencia": estado,
            "resoluciones": " | ".join(data["resoluciones"]),
        })

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(filas[0].keys()) if filas else [])
        writer.writeheader()
        writer.writerows(sorted(filas, key=lambda r: (r["estado_secuencia"], r["codigo_alumno"])))

    observados = sum(1 for f in filas if f["estado_secuencia"] == "Observado")
    print(f"Alumnos con codigo: {len(filas)}")
    print(f"PDFs sin codigo: {sin_codigo}")
    print(f"Secuencia OK: {len(filas) - observados}")
    print(f"Secuencia observada: {observados}")
    print(f"Reporte: {out}")


def comando_limpiar_base(args) -> None:
    if args.confirmar != "BORRAR_BASE_DEV_2026":
        raise RuntimeError("Para limpiar usa: --confirmar BORRAR_BASE_DEV_2026")
    db = SessionLocal()
    try:
        db.query(models.TicketAdjunto).delete()
        db.query(models.HistorialMovimiento).delete()
        db.query(models.AceptacionDictaminante).delete()
        db.query(models.AsignacionTesis).delete()
        db.query(models.ResolucionFirma).delete()
        db.query(models.RevisionTesis).delete()
        if hasattr(models, "ResolucionDocumento"):
            db.query(models.ResolucionDocumento).delete()
        db.query(models.TicketOsticket).update({"id_expediente": None, "estado_scraping": "Pendiente_Descarga"})
        db.query(models.ExpedienteTesis).delete()
        db.query(models.Docente).filter(models.Docente.dni.like("PEN-%")).delete()
        if args.aplicar:
            db.commit()
        else:
            db.rollback()
        print(f"Limpieza {'aplicada' if args.aplicar else 'simulada'}")
    finally:
        db.close()


def build_parser():
    parser = argparse.ArgumentParser(description="Pipeline resoluciones PDF EPG-UAC")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("inventario")
    p.add_argument("--zip", default=str(DEFAULT_ZIP))
    p.add_argument("--out", default=str(DEFAULT_OUT))
    p.set_defaults(func=comando_inventario)

    p = sub.add_parser("extraer")
    p.add_argument("--zip", default=str(DEFAULT_ZIP))
    p.add_argument("--out", default=str(DEFAULT_OUT))
    p.add_argument("--paginas", type=int, default=4)
    p.add_argument("--limite", type=int, default=0)
    p.add_argument("--excel-resoluciones", default=str(DEFAULT_RESOLUCIONES_EXCEL))
    p.set_defaults(func=comando_extraer)

    p = sub.add_parser("extraer-directorio", help="Extrae PDFs descargados desde una carpeta, por ejemplo Google Drive")
    p.add_argument("--directorio", required=True)
    p.add_argument("--out", default=str(ROOT / "data" / "drive_resoluciones" / "extraccion"))
    p.add_argument("--paginas", type=int, default=6)
    p.add_argument("--limite", type=int, default=0)
    p.add_argument("--excel-resoluciones", default=str(DEFAULT_RESOLUCIONES_EXCEL))
    p.set_defaults(func=comando_extraer_directorio)

    p = sub.add_parser("staging")
    p.add_argument("--jsonl", default=str(DEFAULT_OUT / "resoluciones_extraidas.jsonl"))
    p.add_argument("--aplicar", action="store_true")
    p.set_defaults(func=comando_staging)

    p = sub.add_parser("importar-docentes")
    p.add_argument("--excel", default=str(DEFAULT_DOCENTES))
    p.add_argument("--aplicar", action="store_true")
    p.set_defaults(func=comando_importar_docentes)

    p = sub.add_parser("importar-expedientes")
    p.add_argument("--aplicar", action="store_true")
    p.set_defaults(func=comando_importar_expedientes)

    p = sub.add_parser("ordenar-archivos")
    p.add_argument("--zip", default=str(DEFAULT_ZIP))
    p.add_argument("--destino", default=str(DEFAULT_OUT / "pdf_ordenados"))
    p.add_argument("--aplicar", action="store_true")
    p.set_defaults(func=comando_ordenar_archivos)

    p = sub.add_parser("validar-secuencia")
    p.add_argument("--jsonl", default=str(DEFAULT_OUT / "resoluciones_extraidas.jsonl"))
    p.add_argument("--out", default=str(DEFAULT_OUT / "reporte_secuencia.csv"))
    p.set_defaults(func=comando_validar_secuencia)

    p = sub.add_parser("limpiar-base-dev")
    p.add_argument("--confirmar", default="")
    p.add_argument("--aplicar", action="store_true")
    p.set_defaults(func=comando_limpiar_base)
    return parser


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    sys.exit(main())
