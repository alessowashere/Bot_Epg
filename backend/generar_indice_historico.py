"""Construye un índice de consulta desde los Excel históricos, sin tocar la BD."""

from __future__ import annotations

import csv
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl


RAIZ = Path(__file__).resolve().parent.parent
CATALOGOS = RAIZ / "data" / "input" / "catalogos"
SALIDA = RAIZ / "data" / "historico" / "referencias_historicas.csv"
ARCHIVOS = (
    "TRAMITES ADMINISTRATIVOS ESTUDIANTES EPG 2026 (1).xlsx",
    "LISTA DE RESOLUCIONES EMITIDAS 2025 (2).xlsx",
)


def normalizar(texto: object) -> str:
    valor = unicodedata.normalize("NFKD", str(texto or ""))
    valor = "".join(letra for letra in valor if not unicodedata.combining(letra))
    return re.sub(r"\s+", " ", valor).strip().upper()


def codigo(valor: object) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", str(valor or "")).upper()


def fecha(valor: object) -> str:
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    return str(valor or "").strip()


def buscar_columna(headers: list[str], palabras: tuple[str, ...]) -> int | None:
    for indice, header in enumerate(headers):
        if any(palabra in header for palabra in palabras):
            return indice
    return None


def inferir_paso(texto: str) -> str:
    texto = normalizar(texto)
    reglas = (
        ("1", ("ASESOR",)), ("2", ("DICTAMEN DE PROYECTO", "DICTAMINANTE")),
        ("3", ("INSCRIPCION", "INSCRIPCION DE PROYECTO")), ("4", ("APTO AL GRADO",)),
        ("5", ("DICTAMEN DE TESIS",)), ("6", ("FECHA Y HORA", "SUSTENTACION")),
        ("7", ("ELEVAR AL VRAC", "ENTREGA DE GRADO")),
    )
    for paso, frases in reglas:
        if any(frase in texto for frase in frases):
            return paso
    return ""


def indice_cabecera(ws) -> tuple[int, list[str]] | None:
    for fila in range(1, 7):
        headers = [normalizar(cell.value) for cell in ws[fila]]
        encontrados = sum(bool(buscar_columna(headers, grupo)) for grupo in (
            ("NOMBRE", "APELLIDO"), ("CODIGO",), ("RESOLUCION", "TRAMITE", "EXPEDIENTE"),
        ))
        if encontrados >= 2:
            return fila, headers
    return None


def extraer_archivo(ruta: Path) -> list[dict[str, str]]:
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    filas: list[dict[str, str]] = []
    for hoja in libro.worksheets:
        cabecera = indice_cabecera(hoja)
        if not cabecera:
            continue
        numero_fila, headers = cabecera
        col_nombre = buscar_columna(headers, ("NOMBRE", "APELLIDO"))
        col_codigo = buscar_columna(headers, ("CODIGO",))
        col_ticket = buscar_columna(headers, ("TRAMITE", "EXPEDIENTE ADMINISTRATIVO", "SOLICITUD"))
        col_resolucion = buscar_columna(headers, ("N° RESOLUCION", "RESOL. N°", "RESOLUCION"))
        col_tipo = buscar_columna(headers, ("TIPO DE RESOLUCION", "TRAMITE"))
        col_fecha = buscar_columna(headers, ("FECHA DE RESOLUCION", "FECHA"))
        for posicion, valores in enumerate(hoja.iter_rows(min_row=numero_fila + 1, values_only=True), start=numero_fila + 1):
            nombre = str(valores[col_nombre] or "").strip() if col_nombre is not None and col_nombre < len(valores) else ""
            cod = codigo(valores[col_codigo]) if col_codigo is not None and col_codigo < len(valores) else ""
            if "@" in nombre or len(normalizar(nombre)) < 6:
                continue
            tipo = str(valores[col_tipo] or "").strip() if col_tipo is not None and col_tipo < len(valores) else ""
            contexto = f"{hoja.title} {tipo}"
            paso = inferir_paso(contexto)
            # El índice es solo para los siete pasos: no se incorporan asuntos
            # institucionales, resoluciones de Consejo ni filas auxiliares.
            if not paso:
                continue
            filas.append({
                "codigo_alumno": cod,
                "nombre_alumno": normalizar(nombre),
                "ticket_o_expediente": str(valores[col_ticket] or "").strip() if col_ticket is not None and col_ticket < len(valores) else "",
                "resolucion": str(valores[col_resolucion] or "").strip() if col_resolucion is not None and col_resolucion < len(valores) else "",
                "fecha": fecha(valores[col_fecha]) if col_fecha is not None and col_fecha < len(valores) else "",
                "tipo": tipo,
                "paso_sugerido": paso,
                "fuente": ruta.name,
                "hoja": hoja.title,
                "fila": str(posicion),
            })
    return filas


def main() -> None:
    registros = [registro for archivo in ARCHIVOS for registro in extraer_archivo(CATALOGOS / archivo)]
    unicos = {(item["codigo_alumno"], item["nombre_alumno"], item["resolucion"], item["fuente"], item["hoja"], item["fila"]): item for item in registros}
    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    with SALIDA.open("w", encoding="utf-8", newline="") as destino:
        campos = ["codigo_alumno", "nombre_alumno", "ticket_o_expediente", "resolucion", "fecha", "tipo", "paso_sugerido", "fuente", "hoja", "fila"]
        escritor = csv.DictWriter(destino, fieldnames=campos)
        escritor.writeheader()
        escritor.writerows(sorted(unicos.values(), key=lambda item: (item["codigo_alumno"], item["nombre_alumno"])))
    print(f"Índice generado: {SALIDA} ({len(unicos)} referencias)")


if __name__ == "__main__":
    main()
