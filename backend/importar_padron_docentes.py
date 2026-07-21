"""Concilia el padrón EPG local sin borrar docentes extraídos históricamente."""
from __future__ import annotations

import argparse, json, re, unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

from database import SessionLocal
import models

BASE = Path("/opt/DOCENTES")
CONSOLIDADO = BASE / "Docentes_por_programa_EPG_UAC_CON_DICTADOS_ACTUALIZADO_v4_con_contactos.xlsx"
GENERAL = BASE / "Relacion de Docentes EPG UAC ACTUALIZADO.xlsx"
REPORTE = Path("/opt/sistema_posgrado/data/reportes/importacion_padron_docentes.json")

PROGRAMAS_UAC = {
    "Maestria": [
        "Maestría en administración de negocios",
        "Maestría en docencia universitaria",
        "Maestría en contabilidad con mención en auditoría y control interno",
        "Maestría en derecho constitucional",
        "Maestría en derecho registral y notarial",
        "Maestría en ingeniería civil con mención en hidráulica y ambiental",
        "Maestría en ingeniería civil con mención en estructuras",
        "Maestría en derecho del trabajo y de la seguridad social",
        "Maestría en ingeniería civil con mención en transportes",
        "Maestría en ciencias estomatológicas",
        "Maestría en enfermería",
        "Maestría en salud comunitaria",
        "Maestría en seguridad industrial y medio ambiente",
        "Maestría en estadística e investigación científica",
        "Maestría en derecho civil y comercial",
    ],
    "Doctorado": [
        "Doctorado en ciencias de la educación",
        "Doctorado en contabilidad",
        "Doctorado en ciencias de la salud",
        "Doctorado en psicología",
        "Doctorado en medio ambiente y desarrollo sostenible",
        "Doctorado en derecho",
        "Doctorado en administración",
    ],
}


def clave(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode().upper()
    texto = re.sub(r"^(DR|DRA|MG|MGT|MTRO|MTRA|MTR|MAG)\.?\s+", "", texto.strip())
    return re.sub(r"[^A-Z0-9]", "", texto)


def dni(valor):
    hallado = re.search(r"(?<!\d)(\d{8})(?!\d)", str(valor or ""))
    return hallado.group(1) if hallado else None


def fecha(valor):
    if isinstance(valor, datetime): return valor.date()
    for formato in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y"):
        try: return datetime.strptime(str(valor).strip(), formato).date()
        except ValueError: pass
    return None


def main(aplicar=False):
    general = openpyxl.load_workbook(GENERAL, read_only=True, data_only=True)["Datos generales"]
    identidad = {}; identidad_por_dni = {}
    for fila in general.iter_rows(min_row=4, values_only=True):
        if not fila[1]: continue
        info = {
            "nombre": str(fila[1]).strip(), "especialidad": fila[2], "dni": dni(fila[3]),
            "direccion": fila[4], "correo": fila[5], "telefono": fila[6], "titulo": fila[7],
            "universidad": fila[8], "maestria": fila[9], "doctorado": fila[11],
        }
        identidad[clave(fila[1])] = info
        if info["dni"]: identidad_por_dni[info["dni"]] = info

    libro = openpyxl.load_workbook(CONSOLIDADO, read_only=True, data_only=True)
    db = SessionLocal(); creados = actualizados = ambiguos = grados = programas = actividades = catalogo = 0
    try:
        programas_catalogo = {}
        for nivel, nombres in PROGRAMAS_UAC.items():
            for nombre in nombres:
                item = db.query(models.ProgramaPosgrado).filter_by(nombre=nombre).first()
                if not item:
                    slug = clave(nombre).lower()
                    ruta = "maestrias" if nivel == "Maestria" else "doctorados"
                    item = models.ProgramaPosgrado(
                        nivel=nivel, nombre=nombre, slug=slug,
                        url_fuente=f"https://uandina.edu.pe/posgrado/programas/{ruta}", activo=True,
                    )
                    db.add(item); db.flush(); catalogo += 1
                else:
                    item.nivel = nivel; item.activo = True; item.fecha_sincronizacion = datetime.utcnow()
                programas_catalogo[clave(nombre)] = item
        por_dni = {d.dni: d for d in db.query(models.Docente).filter(models.Docente.dni.isnot(None)).all() if d.dni}
        por_nombre = defaultdict(list)
        for d in db.query(models.Docente).all(): por_nombre[clave(d.nombre_completo)].append(d)
        canonicos = {}
        for fila in libro["Todos"].iter_rows(min_row=2, values_only=True):
            if not fila[0]: continue
            k = clave(fila[0]); documento_excel = dni(fila[8]); info = identidad.get(k) or identidad_por_dni.get(documento_excel) or {}; documento = info.get("dni") or documento_excel
            docente = por_dni.get(documento) if documento else None
            if not docente and len(por_nombre[k]) == 1: docente = por_nombre[k][0]
            if not docente:
                if len(por_nombre[k]) > 1: ambiguos += 1
                docente = models.Docente(nombre_completo=info.get("nombre") or str(fila[0]).strip(), dni=documento, tipo_contrato="Semestral", estado="Activo", max_tesis_permitidas=5)
                db.add(docente); db.flush(); creados += 1
                if documento:
                    por_dni[documento] = docente
                por_nombre[k].append(docente)
            else: actualizados += 1
            correo = str(info.get("correo") or fila[9] or "").strip() or None
            docente.dni = docente.dni or documento
            docente.correo = correo or docente.correo
            docente.correo_institucional = correo if correo and correo.lower().endswith("@uandina.edu.pe") else docente.correo_institucional
            docente.correo_personal = correo if correo and not correo.lower().endswith("@uandina.edu.pe") else docente.correo_personal
            docente.telefono = str(info.get("telefono") or fila[10] or "").strip() or docente.telefono
            docente.direccion = str(info.get("direccion") or "").strip() or docente.direccion
            docente.especialidad = str(info.get("especialidad") or "").strip() or docente.especialidad
            docente.titulo_profesional = str(info.get("titulo") or fila[2] or "").strip() or docente.titulo_profesional
            docente.universidad_procedencia = str(info.get("universidad") or fila[1] or "").strip() or docente.universidad_procedencia
            docente.condicion_laboral = str(fila[5] or "").strip() or docente.condicion_laboral
            docente.estado_verificacion = "Padron_EPG"
            docente.fuente_actualizacion = CONSOLIDADO.name
            canonicos[k] = docente

        grados_vistos = {
            (g.id_docente, g.tipo, clave(g.denominacion), g.fecha_diploma)
            for g in db.query(models.DocenteGrado).all()
        }
        for fila in libro["SUNEDU_DETALLE"].iter_rows(min_row=2, values_only=True):
            if not fila[0] or not fila[3]: continue
            docente = canonicos.get(clave(fila[0])); documento = dni(fila[2])
            if not docente and documento: docente = por_dni.get(documento)
            if not docente: continue
            tipo_grado, denominacion, fecha_grado = str(fila[4] or "OTRO"), str(fila[3]).strip(), fecha(fila[7])
            llave_grado = (docente.id_docente, tipo_grado, clave(denominacion), fecha_grado)
            if llave_grado not in grados_vistos:
                db.add(models.DocenteGrado(id_docente=docente.id_docente,tipo=tipo_grado,denominacion=denominacion,universidad=str(fila[5] or "").strip() or None,pais=str(fila[6] or "").strip() or None,fecha_diploma=fecha_grado,fuente="SUNEDU_DETALLE",verificado=True)); grados += 1
                grados_vistos.add(llave_grado)

        programas_vistos = {
            (p.id_docente, p.nivel, clave(p.programa))
            for p in db.query(models.DocentePrograma).all()
        }
        for hoja in libro.sheetnames:
            if not (hoja.startswith("Doc -") or hoja.startswith("Mst -")): continue
            nivel = "Doctorado" if hoja.startswith("Doc -") else "Maestria"
            programa = hoja.split("-", 1)[1].strip()
            for fila in libro[hoja].iter_rows(min_row=2, values_only=True):
                if not fila[0]: continue
                docente = canonicos.get(clave(fila[0]));
                if not docente: continue
                llave_programa = (docente.id_docente, nivel, clave(programa))
                if llave_programa not in programas_vistos:
                    db.add(models.DocentePrograma(id_docente=docente.id_docente,nivel=nivel,programa=programa,tipo_vinculo="Afinidad",especialidad=str(fila[2] or "").strip() or None,estado="Propuesto",fuente=CONSOLIDADO.name)); programas += 1
                    programas_vistos.add(llave_programa)
                else:
                    existente = db.query(models.DocentePrograma).filter_by(id_docente=docente.id_docente,nivel=nivel,programa=programa).first()
                    if existente: existente.tipo_vinculo = "Afinidad"

        if aplicar:
            db.query(models.DocenteActividad).filter(
                models.DocenteActividad.programa.in_(["Programacion EPG", "Programación académica EPG"])
            ).delete(synchronize_session=False)
            db.flush()
        actividad_por_docente_periodo = defaultdict(int)
        for fila in libro["DICTADOS_POR_MES"].iter_rows(min_row=2, values_only=True):
            if not fila[0] or not fila[3]: continue
            docente = canonicos.get(clave(fila[3]));
            if not docente: continue
            periodo_academico = str(fila[2] or "").strip() or None
            mes = str(fila[1] or "").strip() or None
            periodo = str(fila[0])
            programa = "Programación académica EPG"
            actividad_por_docente_periodo[(docente.id_docente, periodo, programa, mes, periodo_academico)] += int(fila[6] or 1)
        for (id_docente, periodo, programa, mes, periodo_academico), cantidad in actividad_por_docente_periodo.items():
            item = db.query(models.DocenteActividad).filter_by(id_docente=id_docente,periodo=periodo,programa=programa).first()
            if item:
                # El padrón es una fotografía autoritativa. Repetir la importación
                # debe conservar el mismo total, no acumularlo otra vez.
                item.registros = cantidad
                item.mes = mes; item.periodo_academico = periodo_academico
                item.tipo_actividad = "Programación docente"; item.detalle = f"{cantidad} participación(es) mensual(es); la fuente no identifica asignatura"
            else:
                db.add(models.DocenteActividad(id_docente=id_docente,periodo=periodo,mes=mes,periodo_academico=periodo_academico,tipo_actividad="Programación docente",programa=programa,detalle=f"{cantidad} participación(es) mensual(es); la fuente no identifica asignatura",registros=cantidad,fuente=CONSOLIDADO.name)); actividades += 1
        if aplicar: db.commit()
        else: db.rollback()
        salida={"modo":"aplicar" if aplicar else "simulacion","creados":creados,"actualizados":actualizados,"nombres_ambiguos":ambiguos,"programas_catalogo_nuevos":catalogo,"grados_nuevos":grados,"afinidades_nuevas":programas,"actividades_nuevas":actividades}
        REPORTE.parent.mkdir(parents=True,exist_ok=True); REPORTE.write_text(json.dumps(salida,ensure_ascii=False,indent=2)+"\n")
        print(json.dumps(salida,ensure_ascii=False,indent=2))
    finally: db.close()


if __name__ == "__main__":
    p=argparse.ArgumentParser();p.add_argument("--aplicar",action="store_true");a=p.parse_args();main(a.aplicar)
