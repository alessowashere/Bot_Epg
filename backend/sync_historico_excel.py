"""
sync_historico_excel.py
───────────────────────
Script de cruce histórico entre los expedientes registrados en la BD
y las filas del Excel de trámites históricos de la EPG-UAC.

Usa Levenshtein (thefuzz) para emparejar nombres con tolerancia tipográfica
y un motor de inferencia de paso basado en palabras clave del Excel.

Uso standalone:
    python sync_historico_excel.py ruta/al/archivo.xlsx --umbral 85 --dry-run

Uso vía API:
    POST /api/admin/sync-historico
"""
import argparse
import sys
from io import BytesIO
from typing import BinaryIO, Optional

# ── Palabras clave para inferir el paso del flujo ──────────────────────────────
INFERENCIA_PASOS = {
    1: ["proyecto", "plan de tesis", "inscripción"],
    2: ["dictaminación", "dictamen", "jurado designado", "resolución de jurados"],
    3: ["borrador", "aprobación de asesor", "revisión de asesor"],
    4: ["sustentación", "fecha de sustentación", "solicitud de sustentación"],
    5: ["acta", "acta de sustentación", "diploma", "certificado"],
    6: ["empaste", "ejemplar", "tesis empastada"],
    7: ["grado", "resolución de grado", "graduado", "diploma académico"],
}


def inferir_paso_desde_texto(texto: str) -> Optional[int]:
    """Analiza el texto del Excel e infiere el id_paso_actual del alumno."""
    texto_lower = texto.lower()
    mejor_paso = None
    mejor_coincidencias = 0

    for id_paso, palabras in INFERENCIA_PASOS.items():
        coincidencias = sum(1 for p in palabras if p in texto_lower)
        if coincidencias > mejor_coincidencias:
            mejor_coincidencias = coincidencias
            mejor_paso = id_paso

    return mejor_paso if mejor_coincidencias > 0 else None


def sync_excel_con_bd(db, archivo: BinaryIO, umbral: int = 85, dry_run: bool = True) -> dict:
    """
    Lee el Excel y cruza nombres con los expedientes de la BD.

    Returns:
        dict con resumen: emparejados, actualizados, sin_match, errores
    """
    try:
        import openpyxl
        from thefuzz import fuzz, process
    except ImportError as e:
        raise RuntimeError(f"Dependencias faltantes: {e}. Instala: pip install thefuzz openpyxl python-Levenshtein")

    # Leer Excel
    contenido = archivo.read() if hasattr(archivo, "read") else archivo
    wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active

    # Leer cabecera
    cabecera = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Detectar columnas relevantes
    col_nombre = _detectar_columna(cabecera, ["nombre", "apellidos", "alumno", "estudiante"])
    col_codigo = _detectar_columna(cabecera, ["codigo", "código", "cu", "expediente"])
    col_tramite = _detectar_columna(cabecera, ["tramite", "trámite", "tipo", "resolución", "estado", "observación"])

    if col_nombre is None:
        raise ValueError(f"No se encontró columna 'nombre' en el Excel. Columnas: {cabecera}")

    # Cargar todos los expedientes de la BD
    import models
    expedientes = db.query(models.ExpedienteTesis).all()
    mapa_nombres = {exp.nombre_alumno.strip().lower(): exp for exp in expedientes}
    mapa_codigos = {str(exp.codigo_alumno).strip(): exp for exp in expedientes if exp.codigo_alumno}
    nombres_lista = list(mapa_nombres.keys())

    resultados = {
        "total_filas_excel": 0,
        "emparejados": 0,
        "actualizados": 0,
        "sin_match": 0,
        "dry_run": dry_run,
        "detalles": [],
    }

    for idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(fila):
            continue

        resultados["total_filas_excel"] += 1

        nombre_excel = str(fila[col_nombre] or "").strip()
        codigo_excel = str(fila[col_codigo] or "").strip() if col_codigo is not None else ""
        texto_tramite = str(fila[col_tramite] or "").strip() if col_tramite is not None else ""

        if not nombre_excel or nombre_excel.lower() in ("none", ""):
            continue

        # 1. Intentar emparejar por código exacto primero
        expediente_match = None
        metodo = ""

        if codigo_excel and codigo_excel in mapa_codigos:
            expediente_match = mapa_codigos[codigo_excel]
            metodo = "codigo_exacto"
            similaridad = 100
        else:
            # 2. Fuzzy matching por nombre
            nombre_lower = nombre_excel.lower()
            match_result = process.extractOne(
                nombre_lower,
                nombres_lista,
                scorer=fuzz.token_sort_ratio,
            )
            if match_result and match_result[1] >= umbral:
                expediente_match = mapa_nombres[match_result[0]]
                metodo = f"fuzzy_{match_result[1]}%"
                similaridad = match_result[1]
            else:
                similaridad = match_result[1] if match_result else 0

        if not expediente_match:
            resultados["sin_match"] += 1
            resultados["detalles"].append({
                "fila_excel": idx,
                "nombre_excel": nombre_excel,
                "similaridad_max": similaridad,
                "accion": "sin_match",
            })
            continue

        resultados["emparejados"] += 1

        # 3. Inferir el paso del flujo desde el texto del trámite
        paso_inferido = inferir_paso_desde_texto(texto_tramite) if texto_tramite else None

        detalle = {
            "fila_excel": idx,
            "nombre_excel": nombre_excel,
            "nombre_bd": expediente_match.nombre_alumno,
            "codigo_bd": expediente_match.codigo_alumno,
            "id_expediente": expediente_match.id_expediente,
            "metodo": metodo,
            "paso_inferido": paso_inferido,
            "paso_actual_bd": expediente_match.id_paso_actual,
            "accion": "sin_cambios",
        }

        # 4. Actualizar BD si hay cambios
        if not dry_run and paso_inferido and paso_inferido != expediente_match.id_paso_actual:
            paso_obj = db.query(models.PasoFlujo).filter(models.PasoFlujo.id_paso == paso_inferido).first()
            if paso_obj:
                from datetime import datetime
                expediente_match.id_paso_actual = paso_inferido
                expediente_match.fecha_ultimo_movimiento = datetime.utcnow()
                db.add(models.HistorialMovimiento(
                    id_expediente=expediente_match.id_expediente,
                    id_paso=paso_inferido,
                    accion="Avanzado",
                    nota=f"Paso inferido desde Excel histórico (fila {idx}): {texto_tramite[:200]}",
                    usuario_nombre="Sistema (Sync Excel)",
                ))
                resultados["actualizados"] += 1
                detalle["accion"] = f"actualizado_a_paso_{paso_inferido}"

        resultados["detalles"].append(detalle)

    if not dry_run:
        db.commit()

    return resultados


def _detectar_columna(cabecera: list, variantes: list) -> Optional[int]:
    """Detecta el índice de la columna que más se parece a alguna variante."""
    for i, col in enumerate(cabecera):
        for variante in variantes:
            if variante in col:
                return i
    return None


# ──────────────────────────────────────────────────────────────────────────────
# CLI para uso desde terminal en el servidor
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync Histórico Excel → BD EPG-UAC")
    parser.add_argument("archivo", help="Ruta al archivo Excel (.xlsx)")
    parser.add_argument("--umbral", type=int, default=85, help="Umbral de similitud Levenshtein (60-100)")
    parser.add_argument("--dry-run", action="store_true", default=True, help="Solo simular, sin modificar BD")
    parser.add_argument("--ejecutar", action="store_true", help="Ejecutar cambios reales en BD")
    args = parser.parse_args()

    dry_run = not args.ejecutar

    # Setup de BD
    sys.path.insert(0, str(__file__))
    from database import SessionLocal
    db = SessionLocal()

    try:
        with open(args.archivo, "rb") as f:
            resultado = sync_excel_con_bd(db, f, args.umbral, dry_run)

        print(f"\n{'='*60}")
        print(f"RESULTADO {'(DRY RUN - sin cambios)' if dry_run else '(EJECUTADO)'}")
        print(f"{'='*60}")
        print(f"Total filas Excel:  {resultado['total_filas_excel']}")
        print(f"Emparejados:        {resultado['emparejados']}")
        print(f"Actualizados en BD: {resultado['actualizados']}")
        print(f"Sin match:          {resultado['sin_match']}")
        print(f"\nPrimeros 20 sin match:")
        sin_match = [d for d in resultado["detalles"] if d["accion"] == "sin_match"]
        for d in sin_match[:20]:
            print(f"  Fila {d['fila_excel']}: '{d['nombre_excel']}' (max sim: {d.get('similaridad_max', '?')}%)")
    finally:
        db.close()
