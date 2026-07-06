import re
import unicodedata
from datetime import datetime

from openpyxl import load_workbook

import models
from extractor import detectar_paso


ALIAS_COLUMNAS = {
    "nombre_alumno": ["nombre", "alumno", "estudiante", "apellidos y nombres", "nombres y apellidos"],
    "codigo_alumno": ["codigo", "codigo alumno", "cod alumno", "cod.", "correo institucional"],
    "grado_postula": ["grado", "programa", "maestria doctorado", "tipo grado"],
    "titulo_tesis": ["titulo", "titulo tesis", "tesis", "proyecto"],
    "id_paso_actual": ["paso", "paso actual", "estado tramite", "tramite", "fase"],
    "resolucion": ["resolucion", "nro resolucion", "numero resolucion", "res."],
}


def normalizar(texto):
    texto = "" if texto is None else str(texto).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto.lower())
    return texto


def resolver_columna(header):
    h = normalizar(header)
    for destino, aliases in ALIAS_COLUMNAS.items():
        if h in aliases or any(alias in h for alias in aliases):
            return destino
    return None


def detectar_fila_encabezados(sheet):
    for row_idx in range(1, min(sheet.max_row, 12) + 1):
        valores = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        reconocidos = [resolver_columna(v) for v in valores if v is not None]
        if len([r for r in reconocidos if r]) >= 2:
            return row_idx
    return 1


def construir_mapa_columnas(sheet, header_row):
    mapa = {}
    for col in range(1, sheet.max_column + 1):
        destino = resolver_columna(sheet.cell(header_row, col).value)
        if destino and destino not in mapa:
            mapa[destino] = col
    return mapa


def texto_celda(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return str(valor).strip()


def normalizar_codigo(valor):
    texto = texto_celda(valor)
    if "@" in texto:
        texto = texto.split("@", 1)[0]
    return texto


def normalizar_grado(valor):
    texto = normalizar(valor)
    if "doctor" in texto:
        return "Doctor"
    return "Maestro"


def normalizar_paso(valor):
    texto = texto_celda(valor)
    match = re.search(r"\b([1-7])\b", texto)
    if match:
        return int(match.group(1))
    sugerido = detectar_paso(texto)
    return sugerido["id_paso"] or 1


def fila_a_dict(sheet, row_idx, mapa):
    data = {}
    for destino, col in mapa.items():
        data[destino] = sheet.cell(row_idx, col).value
    return data


def importar_excel_expedientes(db, file_obj, usuario_nombre="Sistema"):
    file_obj.seek(0)
    workbook = load_workbook(file_obj, data_only=True, read_only=True)

    creados = 0
    actualizados = 0
    omitidos = 0
    errores = []

    for sheet in workbook.worksheets:
        header_row = detectar_fila_encabezados(sheet)
        mapa = construir_mapa_columnas(sheet, header_row)

        requeridas = {"nombre_alumno", "codigo_alumno"}
        if not requeridas.issubset(set(mapa.keys())):
            errores.append({"hoja": sheet.title, "fila": header_row, "error": "No se detectaron columnas de nombre y codigo"})
            continue

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            raw = fila_a_dict(sheet, row_idx, mapa)
            nombre = texto_celda(raw.get("nombre_alumno"))
            codigo = normalizar_codigo(raw.get("codigo_alumno"))
            if not nombre and not codigo:
                omitidos += 1
                continue
            if not nombre or not codigo:
                errores.append({"hoja": sheet.title, "fila": row_idx, "error": "Fila sin nombre o codigo"})
                continue

            grado = normalizar_grado(raw.get("grado_postula"))
            paso = normalizar_paso(raw.get("id_paso_actual"))
            titulo = texto_celda(raw.get("titulo_tesis")) or None
            resolucion = texto_celda(raw.get("resolucion")) or None

            expediente = (
                db.query(models.ExpedienteTesis)
                .filter(models.ExpedienteTesis.codigo_alumno == codigo)
                .order_by(models.ExpedienteTesis.id_expediente.desc())
                .first()
            )

            if expediente:
                expediente.nombre_alumno = nombre
                expediente.grado_postula = grado
                expediente.id_paso_actual = paso
                if titulo:
                    expediente.titulo_tesis = titulo
                expediente.fecha_ultimo_movimiento = datetime.utcnow()
                actualizados += 1
            else:
                expediente = models.ExpedienteTesis(
                    codigo_alumno=codigo,
                    nombre_alumno=nombre,
                    grado_postula=grado,
                    titulo_tesis=titulo,
                    id_paso_actual=paso,
                    estado_expediente="En Proceso",
                )
                db.add(expediente)
                db.flush()
                db.add(
                    models.HistorialMovimiento(
                        id_expediente=expediente.id_expediente,
                        id_paso=paso,
                        accion="Creado",
                        nota=f"Importado desde Excel: {sheet.title}",
                        usuario_nombre=usuario_nombre,
                    )
                )
                creados += 1

            if resolucion:
                existe_res = (
                    db.query(models.ResolucionFirma)
                    .filter(
                        models.ResolucionFirma.id_expediente == expediente.id_expediente,
                        models.ResolucionFirma.tipo_documento == resolucion,
                    )
                    .first()
                )
                if not existe_res:
                    db.add(
                        models.ResolucionFirma(
                            id_expediente=expediente.id_expediente,
                            id_paso_asociado=paso,
                            tipo_documento=resolucion,
                            estado_firma="Firmado",
                            fecha_firma=datetime.utcnow(),
                        )
                    )

        db.commit()

    return {
        "status": "ok",
        "creados": creados,
        "actualizados": actualizados,
        "omitidos": omitidos,
        "errores": errores[:50],
        "errores_total": len(errores),
    }
